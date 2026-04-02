"""
Unit Tests for Excel Import Functionality
测试 Excel 导入功能
"""
import pytest
from io import BytesIO
from openpyxl import Workbook
from app.services.excel_service import ExcelService


class TestExcelParsing:
    """测试 Excel 文件解析"""
    
    def test_parse_valid_excel_file(self):
        """测试解析有效的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头（第1行）
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明行（第2行）
        for idx in range(1, 11):
            ws.cell(row=2, column=idx, value="说明文本")
        
        # 添加示例行（第3行）
        for idx in range(1, 11):
            ws.cell(row=3, column=idx, value="示例数据")
        
        # 添加有效数据行（第4行）
        ws.cell(row=4, column=1, value="192.168.1.10")
        ws.cell(row=4, column=2, value="测试网段")
        ws.cell(row=4, column=3, value="used")
        ws.cell(row=4, column=4, value="测试设备")
        ws.cell(row=4, column=5, value="AA:BB:CC:DD:EE:FF")
        ws.cell(row=4, column=6, value="服务器")
        ws.cell(row=4, column=7, value="张三")
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) == 0, f"Should have no errors, but got: {errors}"
        assert len(valid_data) == 1
        assert valid_data[0]['ip_address'] == "192.168.1.10"
        assert valid_data[0]['segment_name'] == "测试网段"
        assert valid_data[0]['status'] == "used"
    
    def test_parse_excel_with_missing_required_fields(self):
        """测试解析缺少必填字段的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明和示例行
        for row in [2, 3]:
            for idx in range(1, 11):
                ws.cell(row=row, column=idx, value="test")
        
        # 添加缺少必填字段的数据行
        ws.cell(row=4, column=1, value="192.168.1.10")
        # 缺少所属网段
        ws.cell(row=4, column=3, value="used")
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) > 0
        assert len(valid_data) == 0
        assert any("所属网段" in str(error) for error in errors)
    
    def test_parse_excel_with_invalid_ip_format(self):
        """测试解析包含无效 IP 格式的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明和示例行
        for row in [2, 3]:
            for idx in range(1, 11):
                ws.cell(row=row, column=idx, value="test")
        
        # 添加无效 IP 地址的数据行
        ws.cell(row=4, column=1, value="999.999.999.999")  # 无效 IP
        ws.cell(row=4, column=2, value="测试网段")
        ws.cell(row=4, column=3, value="used")
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) > 0
        assert len(valid_data) == 0
        assert any("IP" in str(error) and "无效" in str(error) for error in errors)
    
    def test_parse_excel_with_invalid_mac_format(self):
        """测试解析包含无效 MAC 格式的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明和示例行
        for row in [2, 3]:
            for idx in range(1, 11):
                ws.cell(row=row, column=idx, value="test")
        
        # 添加无效 MAC 地址的数据行
        ws.cell(row=4, column=1, value="192.168.1.10")
        ws.cell(row=4, column=2, value="测试网段")
        ws.cell(row=4, column=3, value="used")
        ws.cell(row=4, column=5, value="INVALID-MAC")  # 无效 MAC
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) > 0
        assert len(valid_data) == 0
        assert any("MAC" in str(error) for error in errors)
    
    def test_parse_excel_with_invalid_status(self):
        """测试解析包含无效状态值的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明和示例行
        for row in [2, 3]:
            for idx in range(1, 11):
                ws.cell(row=row, column=idx, value="test")
        
        # 添加无效状态的数据行
        ws.cell(row=4, column=1, value="192.168.1.10")
        ws.cell(row=4, column=2, value="测试网段")
        ws.cell(row=4, column=3, value="invalid_status")  # 无效状态
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) > 0
        assert len(valid_data) == 0
        assert any("状态" in str(error) for error in errors)
    
    def test_parse_excel_with_empty_rows(self):
        """测试解析包含空行的 Excel 文件"""
        # 创建测试 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        
        # 添加表头
        headers = ["IP地址 *", "所属网段 *", "状态 *", "设备名称", "设备MAC地址", 
                   "设备类型", "责任人", "部门", "位置", "备注"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        # 添加说明和示例行
        for row in [2, 3]:
            for idx in range(1, 11):
                ws.cell(row=row, column=idx, value="test")
        
        # 添加有效数据行
        ws.cell(row=4, column=1, value="192.168.1.10")
        ws.cell(row=4, column=2, value="测试网段")
        ws.cell(row=4, column=3, value="used")
        
        # 第5行为空行
        
        # 添加另一个有效数据行
        ws.cell(row=6, column=1, value="192.168.1.11")
        ws.cell(row=6, column=2, value="测试网段")
        ws.cell(row=6, column=3, value="available")
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果 - 空行应该被跳过
        assert len(errors) == 0
        assert len(valid_data) == 2
    
    def test_parse_excel_with_missing_sheet(self):
        """测试解析缺少'数据导入'工作表的 Excel 文件"""
        # 创建测试 Excel 文件（没有"数据导入"工作表）
        wb = Workbook()
        ws = wb.active
        ws.title = "错误的工作表名"
        
        # 保存到字节流
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 解析文件
        valid_data, errors = ExcelService.parse_import_file(excel_file.read())
        
        # 验证结果
        assert len(errors) > 0
        assert len(valid_data) == 0
        assert any("数据导入" in str(error) for error in errors)


class TestErrorReportGeneration:
    """测试错误报告生成"""
    
    def test_generate_error_report_with_no_errors(self):
        """测试生成无错误的报告"""
        errors = []
        report = ExcelService.generate_error_report(errors)
        
        assert report['has_errors'] is False
        assert report['total_errors'] == 0
        assert len(report['errors']) == 0
    
    def test_generate_error_report_with_single_error(self):
        """测试生成单个错误的报告"""
        errors = [{
            "row": 4,
            "field": "IP地址",
            "value": "999.999.999.999",
            "error": "无效的 IP 地址格式"
        }]
        report = ExcelService.generate_error_report(errors)
        
        assert report['has_errors'] is True
        assert report['total_errors'] == 1
        assert report['affected_rows'] == 1
        assert len(report['errors']) == 1
    
    def test_generate_error_report_with_multiple_errors_same_row(self):
        """测试生成同一行多个错误的报告"""
        errors = [
            {
                "row": 4,
                "field": "IP地址",
                "value": "999.999.999.999",
                "error": "无效的 IP 地址格式"
            },
            {
                "row": 4,
                "field": "MAC地址",
                "value": "INVALID",
                "error": "无效的 MAC 地址格式"
            }
        ]
        report = ExcelService.generate_error_report(errors)
        
        assert report['has_errors'] is True
        assert report['total_errors'] == 2
        assert report['affected_rows'] == 1
        assert len(report['errors']) == 1
        assert report['errors'][0]['error_count'] == 2
    
    def test_generate_error_report_with_multiple_rows(self):
        """测试生成多行错误的报告"""
        errors = [
            {
                "row": 4,
                "field": "IP地址",
                "value": "999.999.999.999",
                "error": "无效的 IP 地址格式"
            },
            {
                "row": 5,
                "field": "状态",
                "value": "invalid",
                "error": "无效的状态值"
            }
        ]
        report = ExcelService.generate_error_report(errors)
        
        assert report['has_errors'] is True
        assert report['total_errors'] == 2
        assert report['affected_rows'] == 2
        assert len(report['errors']) == 2


class TestFieldValidation:
    """测试字段验证"""
    
    def test_validate_ip_address_field(self):
        """测试 IP 地址字段验证"""
        # 有效 IP
        errors = ExcelService._validate_field("ip_address", "192.168.1.10", "IP地址", 4)
        assert len(errors) == 0
        
        # 无效 IP
        errors = ExcelService._validate_field("ip_address", "999.999.999.999", "IP地址", 4)
        assert len(errors) > 0
    
    def test_validate_status_field(self):
        """测试状态字段验证"""
        # 有效状态
        for status in ["available", "used", "reserved"]:
            errors = ExcelService._validate_field("status", status, "状态", 4)
            assert len(errors) == 0
        
        # 无效状态
        errors = ExcelService._validate_field("status", "invalid", "状态", 4)
        assert len(errors) > 0
    
    def test_validate_mac_address_field(self):
        """测试 MAC 地址字段验证"""
        # 有效 MAC
        for mac in ["AA:BB:CC:DD:EE:FF", "aa-bb-cc-dd-ee-ff"]:
            errors = ExcelService._validate_field("device_mac", mac, "设备MAC地址", 4)
            assert len(errors) == 0
        
        # 无效 MAC
        errors = ExcelService._validate_field("device_mac", "INVALID", "设备MAC地址", 4)
        assert len(errors) > 0
    
    def test_validate_text_length(self):
        """测试文本长度验证"""
        # 正常长度
        errors = ExcelService._validate_field("segment_name", "测试网段", "所属网段", 4)
        assert len(errors) == 0
        
        # 超长文本
        long_text = "A" * 101
        errors = ExcelService._validate_field("segment_name", long_text, "所属网段", 4)
        assert len(errors) > 0
    
    def test_validate_special_characters(self):
        """测试特殊字符验证"""
        # 包含特殊字符
        for char in ['<', '>', '&', "'", '"', '\\']:
            text = f"测试{char}文本"
            errors = ExcelService._validate_field("segment_name", text, "所属网段", 4)
            assert len(errors) > 0, f"Should reject special character: {char}"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
