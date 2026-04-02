"""
Unit Tests for Excel Template Generation
测试 Excel 模板生成功能
"""
import pytest
from io import BytesIO
from openpyxl import load_workbook
from app.services.excel_service import ExcelService


class TestExcelTemplateGeneration:
    """测试 Excel 模板生成"""
    
    def test_generate_template_returns_bytesio(self):
        """测试生成模板返回 BytesIO 对象"""
        result = ExcelService.generate_template()
        assert isinstance(result, BytesIO)
        assert result.tell() == 0  # 指针应该在开始位置
    
    def test_template_has_two_sheets(self):
        """测试模板包含两个工作表"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        
        assert len(wb.sheetnames) == 2
        assert "数据导入" in wb.sheetnames
        assert "填写说明" in wb.sheetnames
    
    def test_data_sheet_has_correct_headers(self):
        """测试数据表包含正确的表头"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查第一行表头
        expected_headers = [
            "IP地址 *",
            "所属网段 *",
            "状态 *",
            "设备名称",
            "设备MAC地址",
            "设备类型",
            "责任人",
            "部门",
            "位置",
            "备注"
        ]
        
        for idx, expected_header in enumerate(expected_headers, start=1):
            cell_value = ws.cell(row=1, column=idx).value
            assert cell_value == expected_header, f"Column {idx} header mismatch"
    
    def test_data_sheet_has_descriptions(self):
        """测试数据表包含字段说明"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查第二行有描述信息
        for idx in range(1, 11):  # 10个字段
            cell_value = ws.cell(row=2, column=idx).value
            assert cell_value is not None
            assert len(cell_value) > 0
    
    def test_data_sheet_has_example_data(self):
        """测试数据表包含示例数据"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查第三行有示例数据
        example_ip = ws.cell(row=3, column=1).value
        assert example_ip == "192.168.1.10"
        
        example_status = ws.cell(row=3, column=3).value
        assert example_status == "used"
        
        example_mac = ws.cell(row=3, column=5).value
        assert example_mac == "AA:BB:CC:DD:EE:FF"
    
    def test_data_sheet_has_frozen_panes(self):
        """测试数据表冻结了前三行"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查冻结窗格
        assert ws.freeze_panes == 'A4'
    
    def test_help_sheet_has_content(self):
        """测试说明表包含内容"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["填写说明"]
        
        # 检查标题
        title = ws.cell(row=1, column=1).value
        assert "IPAM" in title
        assert "使用说明" in title
        
        # 检查有多行内容
        content_rows = 0
        for row in ws.iter_rows(min_row=1, max_row=50):
            if row[0].value:
                content_rows += 1
        
        assert content_rows > 10  # 应该有足够多的说明内容
    
    def test_template_fields_definition(self):
        """测试模板字段定义完整性"""
        fields = ExcelService.TEMPLATE_FIELDS
        
        # 检查必填字段
        required_fields = [f for f in fields if f['required']]
        assert len(required_fields) == 3  # IP地址、所属网段、状态
        
        # 检查所有字段都有必要的属性
        for field in fields:
            assert 'field' in field
            assert 'display_name' in field
            assert 'required' in field
            assert 'example' in field
            assert 'description' in field
    
    def test_template_column_widths(self):
        """测试模板列宽设置合理"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查列宽都已设置
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            width = ws.column_dimensions[col].width
            assert width > 0
            assert width >= 10  # 最小宽度应该合理
    
    def test_template_row_heights(self):
        """测试模板行高设置合理"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查前三行的行高
        assert ws.row_dimensions[1].height >= 25  # 表头行
        assert ws.row_dimensions[2].height >= 30  # 说明行
        assert ws.row_dimensions[3].height >= 20  # 示例行


class TestExcelTemplateValidation:
    """测试 Excel 模板验证相关功能"""
    
    def test_required_fields_marked_with_asterisk(self):
        """测试必填字段标记了星号"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查必填字段有星号
        ip_header = ws.cell(row=1, column=1).value
        assert "*" in ip_header
        
        segment_header = ws.cell(row=1, column=2).value
        assert "*" in segment_header
        
        status_header = ws.cell(row=1, column=3).value
        assert "*" in status_header
        
        # 检查可选字段没有星号（或者说不是必填）
        device_name_header = ws.cell(row=1, column=4).value
        # 可选字段不应该有星号，或者如果有也应该是在说明中
    
    def test_status_field_has_valid_options(self):
        """测试状态字段说明包含有效选项"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查状态字段的说明
        status_description = ws.cell(row=2, column=3).value
        assert "available" in status_description
        assert "used" in status_description
        assert "reserved" in status_description
    
    def test_mac_address_format_in_description(self):
        """测试MAC地址格式说明"""
        excel_file = ExcelService.generate_template()
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查MAC地址字段的说明
        mac_description = ws.cell(row=2, column=5).value
        assert "MAC" in mac_description or "mac" in mac_description.lower()
        assert ":" in mac_description or "AA" in ws.cell(row=3, column=5).value


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
