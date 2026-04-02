"""
Integration Tests for Import/Export API
测试 Excel 导入导出 API 端点
"""
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from io import BytesIO

from app.main import app
from app.core.database import get_db
from app.models.user import User
from app.core.security import get_password_hash


@pytest.fixture
def client():
    """创建测试客户端"""
    return TestClient(app)


@pytest.fixture
def test_user(client):
    """创建测试用户并返回认证token"""
    # 这里假设数据库已经初始化并有默认管理员账户
    # 使用默认管理员登录
    response = client.post(
        "/api/v1/auth/login",
        json={
            "username": "admin",
            "password": "admin123"
        }
    )
    
    if response.status_code == 200:
        return response.json()["data"]["access_token"]
    
    # 如果默认管理员不存在，返回None
    return None


class TestExcelTemplateDownload:
    """测试 Excel 模板下载功能"""
    
    def test_download_template_without_auth(self, client):
        """测试未认证用户无法下载模板"""
        response = client.get("/api/v1/import-export/template")
        assert response.status_code == 401
    
    def test_download_template_with_auth(self, client, test_user):
        """测试认证用户可以下载模板"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "Content-Disposition" in response.headers
        assert "IPAM_Import_Template" in response.headers["Content-Disposition"]
    
    def test_downloaded_template_is_valid_excel(self, client, test_user):
        """测试下载的模板是有效的 Excel 文件"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        assert response.status_code == 200
        
        # 尝试加载 Excel 文件
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        # 验证工作表
        assert "数据导入" in wb.sheetnames
        assert "填写说明" in wb.sheetnames
    
    def test_downloaded_template_has_correct_structure(self, client, test_user):
        """测试下载的模板具有正确的结构"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        assert response.status_code == 200
        
        # 加载 Excel 文件
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 验证表头
        assert ws.cell(row=1, column=1).value == "IP地址 *"
        assert ws.cell(row=1, column=2).value == "所属网段 *"
        assert ws.cell(row=1, column=3).value == "状态 *"
        
        # 验证示例数据
        assert ws.cell(row=3, column=1).value == "192.168.1.10"
        assert ws.cell(row=3, column=3).value == "used"
    
    def test_template_filename_contains_timestamp(self, client, test_user):
        """测试模板文件名包含时间戳"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        assert response.status_code == 200
        
        content_disposition = response.headers["Content-Disposition"]
        assert "IPAM_Import_Template_" in content_disposition
        assert ".xlsx" in content_disposition
        
        # 验证文件名格式包含日期时间（格式：YYYYMMDD_HHMMSS）
        import re
        pattern = r"IPAM_Import_Template_\d{8}_\d{6}\.xlsx"
        assert re.search(pattern, content_disposition) is not None


class TestExcelTemplateContent:
    """测试 Excel 模板内容"""
    
    def test_template_has_all_required_fields(self, client, test_user):
        """测试模板包含所有必填字段"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查必填字段标记
        required_fields = []
        for col in range(1, 11):
            header = ws.cell(row=1, column=col).value
            if "*" in header:
                required_fields.append(header)
        
        assert len(required_fields) == 3  # IP地址、所属网段、状态
    
    def test_template_has_field_descriptions(self, client, test_user):
        """测试模板包含字段说明"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["数据导入"]
        
        # 检查第二行的描述
        for col in range(1, 11):
            description = ws.cell(row=2, column=col).value
            assert description is not None
            assert len(description) > 5  # 描述应该有实际内容
    
    def test_help_sheet_has_instructions(self, client, test_user):
        """测试说明表包含使用说明"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/template", headers=headers)
        
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["填写说明"]
        
        # 检查标题
        title = ws.cell(row=1, column=1).value
        assert "IPAM" in title
        assert "使用说明" in title
        
        # 检查有足够的内容行
        content_rows = 0
        for row in range(1, 100):
            if ws.cell(row=row, column=1).value:
                content_rows += 1
        
        assert content_rows > 15  # 应该有足够的说明内容


class TestExcelExport:
    """测试 Excel 导出功能"""
    
    def test_export_without_auth(self, client):
        """测试未认证用户无法导出数据"""
        response = client.get("/api/v1/import-export/export")
        assert response.status_code == 401
    
    def test_export_all_data(self, client, test_user):
        """测试导出所有数据"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export", headers=headers)
        
        # 如果没有数据，应该返回404
        if response.status_code == 404:
            pytest.skip("No data available for export")
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "Content-Disposition" in response.headers
        assert "IPAM_Export" in response.headers["Content-Disposition"]
    
    def test_export_with_segment_filter(self, client, test_user):
        """测试按网段筛选导出"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export?segment_id=1", headers=headers)
        
        # 如果没有数据，应该返回404
        if response.status_code == 404:
            pytest.skip("No data available for export with segment filter")
        
        assert response.status_code == 200
        assert "IPAM_Export_filtered" in response.headers["Content-Disposition"]
    
    def test_export_with_status_filter(self, client, test_user):
        """测试按状态筛选导出"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export?status=used", headers=headers)
        
        # 如果没有数据，应该返回404
        if response.status_code == 404:
            pytest.skip("No data available for export with status filter")
        
        assert response.status_code == 200
    
    def test_export_with_invalid_status(self, client, test_user):
        """测试使用无效状态值导出"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export?status=invalid", headers=headers)
        
        assert response.status_code == 400
        assert "无效的状态值" in response.json()["detail"]
    
    def test_export_with_device_filter(self, client, test_user):
        """测试按设备筛选导出"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export?device_id=1", headers=headers)
        
        # 如果没有数据，应该返回404
        if response.status_code == 404:
            pytest.skip("No data available for export with device filter")
        
        assert response.status_code == 200
    
    def test_export_with_multiple_filters(self, client, test_user):
        """测试使用多个筛选条件导出"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get(
            "/api/v1/import-export/export?segment_id=1&status=used",
            headers=headers
        )
        
        # 如果没有数据，应该返回404
        if response.status_code == 404:
            pytest.skip("No data available for export with multiple filters")
        
        assert response.status_code == 200
        assert "IPAM_Export_filtered" in response.headers["Content-Disposition"]
    
    def test_exported_file_is_valid_excel(self, client, test_user):
        """测试导出的文件是有效的 Excel 文件"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export", headers=headers)
        
        if response.status_code == 404:
            pytest.skip("No data available for export")
        
        assert response.status_code == 200
        
        # 尝试加载 Excel 文件
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        
        # 验证工作表
        assert "IPAM数据导出" in wb.sheetnames
    
    def test_exported_file_has_correct_headers(self, client, test_user):
        """测试导出文件包含正确的表头"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export", headers=headers)
        
        if response.status_code == 404:
            pytest.skip("No data available for export")
        
        assert response.status_code == 200
        
        # 加载 Excel 文件
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["IPAM数据导出"]
        
        # 验证表头（包含所有字段）
        expected_headers = [
            "IP地址", "所属网段", "状态", "设备名称", "设备MAC地址",
            "设备类型", "责任人", "部门", "位置", "备注",
            "分配时间", "分配人", "最后扫描时间", "在线状态"
        ]
        
        for idx, expected_header in enumerate(expected_headers, start=1):
            actual_header = ws.cell(row=1, column=idx).value
            assert actual_header == expected_header, f"Column {idx}: expected '{expected_header}', got '{actual_header}'"
    
    def test_exported_file_has_data_rows(self, client, test_user):
        """测试导出文件包含数据行"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export", headers=headers)
        
        if response.status_code == 404:
            pytest.skip("No data available for export")
        
        assert response.status_code == 200
        
        # 加载 Excel 文件
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb["IPAM数据导出"]
        
        # 检查是否有数据行（至少有表头 + 1行数据）
        assert ws.max_row >= 2, "Export file should have at least one data row"
    
    def test_export_filename_contains_timestamp(self, client, test_user):
        """测试导出文件名包含时间戳"""
        if test_user is None:
            pytest.skip("No test user available")
        
        headers = {"Authorization": f"Bearer {test_user}"}
        response = client.get("/api/v1/import-export/export", headers=headers)
        
        if response.status_code == 404:
            pytest.skip("No data available for export")
        
        assert response.status_code == 200
        
        content_disposition = response.headers["Content-Disposition"]
        assert "IPAM_Export" in content_disposition
        assert ".xlsx" in content_disposition
        
        # 验证文件名格式包含日期时间（格式：YYYYMMDD_HHMMSS）
        import re
        pattern = r"IPAM_Export_(all|filtered)_\d{8}_\d{6}\.xlsx"
        assert re.search(pattern, content_disposition) is not None


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

