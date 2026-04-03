"""
Excel Service - Excel 导入导出服务
处理 Excel 模板生成、数据导入和导出功能
"""
import io
import re
import ipaddress
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.utils.device_utils import validate_mac_address, normalize_mac_address


class ExcelService:
    """Excel 服务类，处理 Excel 相关操作"""

    # Excel 模板字段定义
    TEMPLATE_FIELDS = [
        {
            "field": "ip_address",
            "display_name": "IP地址",
            "required": True,
            "example": "192.168.1.10",
            "description": "IP地址，格式如：192.168.1.10"
        },
        {
            "field": "segment_name",
            "display_name": "所属网段",
            "required": True,
            "example": "办公网段",
            "description": "IP地址所属的网段名称"
        },
        {
            "field": "status",
            "display_name": "状态",
            "required": True,
            "example": "used",
            "description": "IP状态：available(空闲)/used(已用)/reserved(保留)"
        },
        {
            "field": "device_name",
            "display_name": "设备名称",
            "required": False,
            "example": "服务器-01",
            "description": "关联的设备名称（如果IP已分配）"
        },
        {
            "field": "device_mac",
            "display_name": "设备MAC地址",
            "required": False,
            "example": "AA:BB:CC:DD:EE:FF",
            "description": "设备MAC地址，格式如：AA:BB:CC:DD:EE:FF"
        },
        {
            "field": "device_type",
            "display_name": "设备类型",
            "required": False,
            "example": "服务器",
            "description": "设备类型：服务器/交换机/路由器/终端等"
        },
        {
            "field": "owner",
            "display_name": "责任人",
            "required": False,
            "example": "张三",
            "description": "设备责任人姓名"
        },
        {
            "field": "department",
            "display_name": "部门",
            "required": False,
            "example": "技术部",
            "description": "设备所属部门"
        },
        {
            "field": "location",
            "display_name": "位置",
            "required": False,
            "example": "机房A-01",
            "description": "设备物理位置"
        },
        {
            "field": "description",
            "display_name": "备注",
            "required": False,
            "example": "Web服务器",
            "description": "设备或IP的备注信息"
        }
    ]

    @staticmethod
    def generate_template() -> io.BytesIO:
        """
        生成 Excel 导入模板

        Returns:
            BytesIO: Excel 文件的字节流
        """
        # 创建工作簿
        wb = Workbook()

        # 创建数据表
        ws_data = wb.active
        ws_data.title = "数据导入"

        # 创建说明表
        ws_help = wb.create_sheet("填写说明")

        # === 填充数据表 ===
        ExcelService._fill_data_sheet(ws_data)

        # === 填充说明表 ===
        ExcelService._fill_help_sheet(ws_help)

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def _fill_data_sheet(ws):
        """填充数据表"""
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        example_font = Font(name='微软雅黑', size=10, color='666666', italic=True)
        example_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 设置列宽
        column_widths = {
            'A': 18,  # IP地址
            'B': 15,  # 所属网段
            'C': 12,  # 状态
            'D': 18,  # 设备名称
            'E': 20,  # 设备MAC地址
            'F': 12,  # 设备类型
            'G': 12,  # 责任人
            'H': 12,  # 部门
            'I': 15,  # 位置
            'J': 25   # 备注
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 第一行：字段显示名称（带必填标记）
        for idx, field in enumerate(ExcelService.TEMPLATE_FIELDS, start=1):
            cell = ws.cell(row=1, column=idx)
            display_name = field['display_name']
            if field['required']:
                display_name += " *"
            cell.value = display_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 第二行：字段说明
        for idx, field in enumerate(ExcelService.TEMPLATE_FIELDS, start=1):
            cell = ws.cell(row=2, column=idx)
            cell.value = field['description']
            cell.font = Font(name='微软雅黑', size=9, color='666666')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border

        # 第三行：示例数据
        for idx, field in enumerate(ExcelService.TEMPLATE_FIELDS, start=1):
            cell = ws.cell(row=3, column=idx)
            cell.value = field['example']
            cell.font = example_font
            cell.fill = example_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border

        # 设置行高
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 25

        # 冻结前三行
        ws.freeze_panes = 'A4'

    @staticmethod
    def _fill_help_sheet(ws):
        """填充说明表"""
        # 标题样式
        title_font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
        section_font = Font(name='微软雅黑', size=12, bold=True)
        normal_font = Font(name='微软雅黑', size=10)

        # 设置列宽
        ws.column_dimensions['A'].width = 80

        # 添加说明内容
        row = 1

        # 标题
        cell = ws.cell(row=row, column=1)
        cell.value = "IPAM 系统数据导入模板使用说明"
        cell.font = title_font
        row += 2

        # 概述
        cell = ws.cell(row=row, column=1)
        cell.value = "一、概述"
        cell.font = section_font
        row += 1

        cell = ws.cell(row=row, column=1)
        cell.value = "本模板用于批量导入 IP 地址和设备信息到 IPAM 系统。请按照以下说明填写数据。"
        cell.font = normal_font
        row += 2

        # 必填字段
        cell = ws.cell(row=row, column=1)
        cell.value = "二、必填字段（标记 * 的字段）"
        cell.font = section_font
        row += 1

        required_fields = [f for f in ExcelService.TEMPLATE_FIELDS if f['required']]
        for field in required_fields:
            cell = ws.cell(row=row, column=1)
            cell.value = f"• {field['display_name']}：{field['description']}"
            cell.font = normal_font
            row += 1
        row += 1

        # 可选字段
        cell = ws.cell(row=row, column=1)
        cell.value = "三、可选字段"
        cell.font = section_font
        row += 1

        optional_fields = [f for f in ExcelService.TEMPLATE_FIELDS if not f['required']]
        for field in optional_fields:
            cell = ws.cell(row=row, column=1)
            cell.value = f"• {field['display_name']}：{field['description']}"
            cell.font = normal_font
            row += 1
        row += 1

        # 数据格式要求
        cell = ws.cell(row=row, column=1)
        cell.value = "四、数据格式要求"
        cell.font = section_font
        row += 1

        format_rules = [
            "1. IP地址格式：必须是有效的 IPv4 地址，如 192.168.1.10",
            "2. 网段名称：必须是系统中已存在的网段名称",
            "3. 状态：只能填写 available（空闲）、used（已用）或 reserved（保留）",
            "4. MAC地址格式：必须是有效的 MAC 地址，如 AA:BB:CC:DD:EE:FF 或 aa-bb-cc-dd-ee-ff",
            "5. 如果 IP 状态为 used（已用），建议填写设备相关信息",
            "6. 设备MAC地址在系统中必须唯一，不能重复",
            "7. 所有文本字段不能包含特殊字符（如 <、>、&、' 等）"
        ]

        for rule in format_rules:
            cell = ws.cell(row=row, column=1)
            cell.value = rule
            cell.font = normal_font
            row += 1
        row += 1

        # 导入流程
        cell = ws.cell(row=row, column=1)
        cell.value = "五、导入流程"
        cell.font = section_font
        row += 1

        import_steps = [
            "1. 删除或保留第3行的示例数据（示例数据不会被导入）",
            "2. 从第4行开始填写实际数据，每行一条记录",
            "3. 填写完成后保存文件",
            "4. 在 IPAM 系统中选择'导入数据'功能",
            "5. 上传填写好的 Excel 文件",
            "6. 系统会自动验证数据格式和完整性",
            "7. 如果验证通过，数据将被导入系统",
            "8. 如果验证失败，系统会返回详细的错误报告"
        ]

        for step in import_steps:
            cell = ws.cell(row=row, column=1)
            cell.value = step
            cell.font = normal_font
            row += 1
        row += 1

        # 注意事项
        cell = ws.cell(row=row, column=1)
        cell.value = "六、注意事项"
        cell.font = section_font
        row += 1

        notes = [
            "• 请勿修改表头（第1行）的字段名称",
            "• 请勿删除或调整列的顺序",
            "• 导入前请确保网段已在系统中创建",
            "• 如果 IP 地址已存在，导入时会跳过该记录",
            "• 如果设备 MAC 地址已存在，会关联到现有设备",
            "• 建议先导入少量数据测试，确认无误后再批量导入",
            "• 导入大量数据时请耐心等待，不要关闭页面"
        ]

        for note in notes:
            cell = ws.cell(row=row, column=1)
            cell.value = note
            cell.font = normal_font
            row += 1
        row += 1

        # 联系方式
        cell = ws.cell(row=row, column=1)
        cell.value = "如有问题，请联系系统管理员。"
        cell.font = Font(name='微软雅黑', size=10, italic=True, color='666666')

        # 设置所有行的自动换行
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    @staticmethod
    def parse_import_file(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        解析上传的 Excel 文件

        Args:
            file_content: Excel 文件的字节内容

        Returns:
            Tuple[List[Dict], List[Dict]]: (有效数据列表, 错误列表)
        """
        try:
            # 加载 Excel 文件
            wb = load_workbook(io.BytesIO(file_content))

            # 检查是否有"数据导入"工作表
            if "数据导入" not in wb.sheetnames:
                return [], [{
                    "row": 0,
                    "field": "file",
                    "value": "",
                    "error": "Excel 文件格式错误：缺少'数据导入'工作表"
                }]

            ws = wb["数据导入"]

            # 验证表头
            header_errors = ExcelService._validate_headers(ws)
            if header_errors:
                return [], header_errors

            # 解析数据行（从第4行开始，跳过表头、说明和示例）
            valid_data = []
            errors = []

            for row_idx in range(4, ws.max_row + 1):
                # 检查是否为空行
                if ExcelService._is_empty_row(ws, row_idx):
                    continue

                # 解析行数据
                row_data, row_errors = ExcelService._parse_row(ws, row_idx)

                if row_errors:
                    errors.extend(row_errors)
                else:
                    valid_data.append(row_data)

            return valid_data, errors

        except Exception as e:
            return [], [{
                "row": 0,
                "field": "file",
                "value": "",
                "error": f"解析 Excel 文件失败: {str(e)}"
            }]

    @staticmethod
    def _validate_headers(ws) -> List[Dict[str, Any]]:
        """验证 Excel 表头是否正确"""
        expected_headers = [field['display_name'] for field in ExcelService.TEMPLATE_FIELDS]
        errors = []

        for idx, expected_header in enumerate(expected_headers, start=1):
            cell_value = ws.cell(row=1, column=idx).value
            # 移除必填标记进行比较
            if cell_value:
                cell_value = cell_value.replace(" *", "").strip()
            expected_header_clean = expected_header.replace(" *", "").strip()

            if cell_value != expected_header_clean:
                errors.append({
                    "row": 1,
                    "field": f"column_{idx}",
                    "value": cell_value or "",
                    "error": f"表头错误：第{idx}列应为'{expected_header_clean}'，实际为'{cell_value}'"
                })

        return errors

    @staticmethod
    def _is_empty_row(ws, row_idx: int) -> bool:
        """检查是否为空行"""
        for col_idx in range(1, 11):  # 检查前10列
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():
                return False
        return True

    @staticmethod
    def _parse_row(ws, row_idx: int) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """
        解析单行数据

        Returns:
            Tuple[Dict, List[Dict]]: (行数据字典, 错误列表)
        """
        row_data = {}
        errors = []

        # 读取所有字段
        for idx, field_def in enumerate(ExcelService.TEMPLATE_FIELDS, start=1):
            field_name = field_def['field']
            cell_value = ws.cell(row=row_idx, column=idx).value

            # 转换为字符串并去除空白
            if cell_value is not None:
                cell_value = str(cell_value).strip()

            # 检查必填字段
            if field_def['required'] and not cell_value:
                errors.append({
                    "row": row_idx,
                    "field": field_def['display_name'],
                    "value": "",
                    "error": f"必填字段'{field_def['display_name']}'不能为空"
                })
                continue

            # 如果是可选字段且为空，跳过
            if not cell_value:
                row_data[field_name] = None
                continue

            # 字段特定验证
            field_errors = ExcelService._validate_field(
                field_name,
                cell_value,
                field_def['display_name'],
                row_idx
            )

            if field_errors:
                errors.extend(field_errors)
            else:
                row_data[field_name] = cell_value

        # 如果有错误，返回空数据
        if errors:
            return {}, errors

        return row_data, []

    @staticmethod
    def _validate_field(field_name: str, value: str, display_name: str, row_idx: int) -> List[Dict[str, Any]]:
        """
        验证单个字段的值

        Returns:
            List[Dict]: 错误列表
        """
        errors = []

        # IP 地址验证
        if field_name == "ip_address":
            try:
                ipaddress.ip_address(value)
            except ValueError:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"无效的 IP 地址格式: {value}"
                })

        # 状态验证
        elif field_name == "status":
            valid_statuses = ["available", "used", "reserved"]
            if value.lower() not in valid_statuses:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"无效的状态值: {value}，必须是 available、used 或 reserved 之一"
                })

        # MAC 地址验证
        elif field_name == "device_mac":
            is_valid, error_msg = validate_mac_address(value)
            if not is_valid:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"无效的 MAC 地址格式: {error_msg}"
                })

        # 文本长度验证
        elif field_name in ["segment_name", "device_name", "owner"]:
            if len(value) > 100:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"字段'{display_name}'长度不能超过100个字符"
                })

        elif field_name in ["device_type", "department"]:
            if len(value) > 50:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"字段'{display_name}'长度不能超过50个字符"
                })

        elif field_name == "location":
            if len(value) > 200:
                errors.append({
                    "row": row_idx,
                    "field": display_name,
                    "value": value,
                    "error": f"字段'{display_name}'长度不能超过200个字符"
                })

        # 检查特殊字符
        if value and re.search(r'[<>&\'"\\]', value):
            errors.append({
                "row": row_idx,
                "field": display_name,
                "value": value,
                "error": f"字段'{display_name}'包含不允许的特殊字符 (<, >, &, ', \", \\)"
            })

        return errors

    @staticmethod
    def generate_error_report(errors: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        生成详细的错误报告

        Args:
            errors: 错误列表

        Returns:
            Dict: 错误报告
        """
        if not errors:
            return {
                "has_errors": False,
                "total_errors": 0,
                "errors": []
            }

        # 按行分组错误
        errors_by_row = {}
        for error in errors:
            row = error['row']
            if row not in errors_by_row:
                errors_by_row[row] = []
            errors_by_row[row].append({
                "field": error['field'],
                "value": error['value'],
                "error": error['error']
            })

        # 生成报告
        error_list = []
        for row, row_errors in sorted(errors_by_row.items()):
            if row == 0:
                # 文件级别错误
                error_list.extend(row_errors)
            else:
                error_list.append({
                    "row": row,
                    "errors": row_errors,
                    "error_count": len(row_errors)
                })

        return {
            "has_errors": True,
            "total_errors": len(errors),
            "affected_rows": len([r for r in errors_by_row.keys() if r > 0]),
            "errors": error_list
        }

    @staticmethod
    def validate_import_data(data: List[Dict[str, Any]], db) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        验证导入数据的业务逻辑

        Args:
            data: 解析后的数据列表
            db: 数据库会话

        Returns:
            Tuple[List[Dict], List[Dict]]: (有效数据列表, 错误列表)
        """
        from app.models.network_segment import NetworkSegment
        from app.models.ip_address import IPAddress
        from app.models.device import Device

        valid_data = []
        errors = []

        # 缓存网段信息
        segments_cache = {}

        for idx, row_data in enumerate(data, start=4):  # 从第4行开始
            row_errors = []

            # 验证网段是否存在
            segment_name = row_data.get('segment_name')
            if segment_name:
                if segment_name not in segments_cache:
                    segment = db.query(NetworkSegment).filter(
                        NetworkSegment.name == segment_name
                    ).first()
                    if segment:
                        segments_cache[segment_name] = segment
                    else:
                        row_errors.append({
                            "row": idx,
                            "field": "所属网段",
                            "value": segment_name,
                            "error": f"网段'{segment_name}'不存在，请先在系统中创建该网段"
                        })

                if segment_name in segments_cache:
                    row_data['segment_id'] = segments_cache[segment_name].id

            # 验证 IP 地址是否属于网段
            if 'segment_id' in row_data and row_data.get('ip_address'):
                segment = segments_cache[segment_name]
                ip_addr = row_data['ip_address']

                # 检查 IP 是否在网段范围内
                try:
                    network = ipaddress.ip_network(f"{segment.network}/{segment.prefix_length}", strict=False)
                    ip_obj = ipaddress.ip_address(ip_addr)

                    if ip_obj not in network:
                        row_errors.append({
                            "row": idx,
                            "field": "IP地址",
                            "value": ip_addr,
                            "error": f"IP地址 {ip_addr} 不在网段 {segment.network}/{segment.prefix_length} 范围内"
                        })
                except Exception as e:
                    row_errors.append({
                        "row": idx,
                        "field": "IP地址",
                        "value": ip_addr,
                        "error": f"验证 IP 地址范围时出错: {str(e)}"
                    })

            # 检查 IP 地址是否已存在
            if row_data.get('ip_address'):
                existing_ip = db.query(IPAddress).filter(
                    IPAddress.ip_address == row_data['ip_address']
                ).first()

                if existing_ip:
                    row_errors.append({
                        "row": idx,
                        "field": "IP地址",
                        "value": row_data['ip_address'],
                        "error": f"IP地址 {row_data['ip_address']} 已存在于系统中"
                    })

            # 标准化 MAC 地址
            if row_data.get('device_mac'):
                row_data['device_mac'] = normalize_mac_address(row_data['device_mac'])

            # 检查状态和设备信息的一致性
            status = row_data.get('status', '').lower()
            has_device_info = any([
                row_data.get('device_name'),
                row_data.get('device_mac'),
                row_data.get('device_type'),
                row_data.get('owner')
            ])

            if status == 'used' and not has_device_info:
                row_errors.append({
                    "row": idx,
                    "field": "状态",
                    "value": status,
                    "error": "IP状态为'used'时，建议填写设备相关信息（至少包含设备名称和责任人）"
                })

            if row_errors:
                errors.extend(row_errors)
            else:
                valid_data.append(row_data)

        return valid_data, errors

    @staticmethod
    def generate_export_file(data: List[Dict[str, Any]]) -> io.BytesIO:
        """
        生成 Excel 导出文件

        Args:
            data: 要导出的数据列表，每个元素是一个包含所有字段的字典

        Returns:
            BytesIO: Excel 文件的字节流
        """
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "IPAM数据导出"

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')

        border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 设置列宽
        column_widths = {
            'A': 18,  # IP地址
            'B': 15,  # 所属网段
            'C': 12,  # 状态
            'D': 18,  # 设备名称
            'E': 20,  # 设备MAC地址
            'F': 12,  # 设备类型
            'G': 12,  # 责任人
            'H': 12,  # 部门
            'I': 15,  # 位置
            'J': 25,  # 备注
            'K': 20,  # 分配时间
            'L': 15,  # 分配人
            'M': 20,  # 最后扫描时间
            'N': 10   # 在线状态
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 写入表头
        headers = [
            "IP地址", "所属网段", "状态", "设备名称", "设备MAC地址",
            "设备类型", "责任人", "部门", "位置", "备注",
            "分配时间", "分配人", "最后扫描时间", "在线状态"
        ]

        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 设置表头行高
        ws.row_dimensions[1].height = 30

        # 写入数据
        for row_idx, row_data in enumerate(data, start=2):
            # IP地址
            cell = ws.cell(row=row_idx, column=1)
            cell.value = row_data.get('ip_address', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 所属网段
            cell = ws.cell(row=row_idx, column=2)
            cell.value = row_data.get('segment_name', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 状态
            cell = ws.cell(row=row_idx, column=3)
            status = row_data.get('status', '')
            # 状态翻译
            status_map = {
                'available': '空闲',
                'used': '已用',
                'reserved': '保留'
            }
            cell.value = status_map.get(status, status)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 根据状态设置背景色
            if status == 'available':
                cell.fill = PatternFill(start_color='E7F4E4', end_color='E7F4E4', fill_type='solid')
            elif status == 'used':
                cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
            elif status == 'reserved':
                cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

            # 设备名称
            cell = ws.cell(row=row_idx, column=4)
            cell.value = row_data.get('device_name', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 设备MAC地址
            cell = ws.cell(row=row_idx, column=5)
            cell.value = row_data.get('device_mac', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 设备类型
            cell = ws.cell(row=row_idx, column=6)
            cell.value = row_data.get('device_type', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 责任人
            cell = ws.cell(row=row_idx, column=7)
            cell.value = row_data.get('owner', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 部门
            cell = ws.cell(row=row_idx, column=8)
            cell.value = row_data.get('department', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 位置
            cell = ws.cell(row=row_idx, column=9)
            cell.value = row_data.get('location', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 备注
            cell = ws.cell(row=row_idx, column=10)
            cell.value = row_data.get('description', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 分配时间
            cell = ws.cell(row=row_idx, column=11)
            allocated_at = row_data.get('allocated_at')
            if allocated_at:
                if isinstance(allocated_at, str):
                    cell.value = allocated_at
                else:
                    cell.value = allocated_at.strftime('%Y-%m-%d %H:%M:%S')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 分配人
            cell = ws.cell(row=row_idx, column=12)
            cell.value = row_data.get('allocated_by_username', '')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 最后扫描时间
            cell = ws.cell(row=row_idx, column=13)
            last_seen = row_data.get('last_seen')
            if last_seen:
                if isinstance(last_seen, str):
                    cell.value = last_seen
                else:
                    cell.value = last_seen.strftime('%Y-%m-%d %H:%M:%S')
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

            # 在线状态
            cell = ws.cell(row=row_idx, column=14)
            is_online = row_data.get('is_online')
            if is_online is not None:
                cell.value = '在线' if is_online else '离线'
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

        # 冻结表头
        ws.freeze_panes = 'A2'

        # 添加筛选器
        ws.auto_filter.ref = ws.dimensions

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output
